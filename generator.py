from __future__ import annotations

import io
import math
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DEFAULT_PLANTS = ["1000", "1100", "2000", "2100"]
DEMAND_PATTERNS = [
    ("stable", 0.35),
    ("volatile", 0.20),
    ("seasonal", 0.15),
    ("intermittent", 0.20),
    ("zero", 0.10),
]
ABC_DISTRIBUTION = [("A", 0.20), ("B", 0.30), ("C", 0.50)]
PRICE_CONTROL_DISTRIBUTION = [("S", 0.70), ("V", 0.30)]
SERVICE_LEVELS = {"A": 0.98, "B": 0.95, "C": 0.90}
Z_VALUES = {"A": 2.05, "B": 1.65, "C": 1.28}
MOVEMENT_TYPES = np.array([201, 261, 281, 601, 641, 643, 645])
MOVEMENT_SIGNS = np.array(["S", "H"])
UOM_CONFIG = {
    "ST": {"decan": 0},
    "PC": {"decan": 0},
    "KG": {"decan": 3},
    "L": {"decan": 3},
}
TEMPLATE_STRUCTURE = {
    "T001": [
        ("Buchungskreis", "BUKRS"),
        ("Währung", "WAERS"),
    ],
    "T001W": [
        ("Werk", "WERKS"),
        ("Bewertungskreis", "BWKEY"),
    ],
    "T006": [
        ("Maßeinheit", "MSEHI"),
        ("Anzahl Dezimalstellen", "DECAN"),
    ],
    "T134": [
        ("Materialart", "MTART"),
        ("Produkttyp (Material / Service / Subscription)", "PROD_TYPE_CODE"),
    ],
    "V438A": [
        ("Dispositionsverfahren", "DISMM"),
        ("Dispositionskennzeichen (aus DISMM abgeleitet)", "DISVF"),
    ],
    "V_399D_E": [
        ("Werk", "WERKS"),
        ("Einkaufsbearbeitungszeit (Werktage)", "BZTEK"),
    ],
    "MARA": [
        ("Materialnummer", "MATNR"),
        ("Materialart", "MTART"),
        ("Löschvormerkung", "LVORM"),
        ("Basismengeneinheit", "MEINS"),
    ],
    "MATDOC": [
        ("Materialnummer", "MATNR"),
        ("Werk", "WERKS"),
        ("Bewegungsart", "BWART"),
        ("Soll-/Haben-Kennzeichen", "SHKZG"),
        ("Menge", "MENGE"),
        ("Buchungsdatum", "BUDAT"),
        ("Materialbelegjahr", "MJAHR"),
        ("Materialbelegnummer", "MBLNR"),
        ("Belegzeile", "ZEILE"),
        ("Stornojahr", "SJAHR"),
        ("Stornobeleg", "SMBLN"),
        ("Stornozeile", "SMBLP"),
    ],
    "MARC": [
        ("Materialnummer", "MATNR"),
        ("Werk", "WERKS"),
        ("Löschvormerkung", "LVORM"),
        ("Dispositionsmerkmal", "DISMM"),
        ("Ist-Sicherheitsbestand", "EISBE"),
        ("Planlieferzeit (Kalendertage)", "PLIFZ"),
        ("Wareneingangsbearbeitungszeit (Werktage)", "WEBAZ"),
        ("Pflegestatus (D/E/L)", "PSTAT"),
        ("ABC-Indikator", "MAABC"),
        ("Wiederbeschaffungszeit (optional)", "WZEIT"),
    ],
    "MBEW": [
        ("Materialnummer", "MATNR"),
        ("Bewertungskreis", "BWKEY"),
        ("Preissteuerung", "VPRSV"),
        ("Standardpreis", "STPRS"),
        ("Gleitender Durchschnittspreis", "VERPR"),
        ("Preiseinheit", "PEINH"),
        ("Löschvormerkung", "LVORM"),
    ],
}


@dataclass(frozen=True)
class MaterialRecord:
    matnr: str
    werks: str
    mtart: str
    meins: str
    demand_pattern: str
    abc_class: str
    lead_time_days: int
    goods_receipt_days: int
    purchasing_processing_days: int
    wzeit_days: int | None
    safety_stock: float | None
    price_control: str
    standard_price: float
    moving_average_price: float
    deletion_flag: str
    plant_deletion_flag: str
    base_monthly_demand: float


def build_rng() -> tuple[np.random.Generator, int]:
    seed = time.time_ns()
    return np.random.default_rng(seed), seed


def weighted_choice(rng: np.random.Generator, options: list[tuple[str, float]]) -> str:
    labels = [label for label, _ in options]
    weights = [weight for _, weight in options]
    return str(rng.choice(labels, p=weights))


def generate_plant_codes(num_plants: int) -> list[str]:
    if num_plants <= len(DEFAULT_PLANTS):
        return DEFAULT_PLANTS[:num_plants]

    plants = DEFAULT_PLANTS.copy()
    next_code = 2200
    while len(plants) < num_plants:
        plants.append(str(next_code))
        next_code += 100
    return plants


def generate_material_id(plant_idx: int, material_idx: int) -> str:
    return f"MAT{plant_idx + 1:02d}{material_idx + 1:04d}"


def company_code(werks: str) -> str:
    return f"{werks[:2]}00"


def generate_price(rng: np.random.Generator) -> float:
    bucket = weighted_choice(rng, [("cheap", 0.30), ("medium", 0.50), ("expensive", 0.20)])
    if bucket == "cheap":
        return round(float(rng.uniform(0.5, 10.0)), 2)
    if bucket == "medium":
        return round(float(rng.uniform(10.0, 200.0)), 2)
    return round(float(rng.uniform(200.0, 5000.0)), 2)


def generate_demand_dates(rng: np.random.Generator, pattern: str, start_date: date, end_date: date) -> list[date]:
    days = max((end_date - start_date).days, 1)
    if pattern == "zero":
        count = int(rng.integers(0, 3))
    elif pattern == "stable":
        count = int(rng.integers(max(12, days // 10), max(20, days // 5)))
    elif pattern == "volatile":
        count = int(rng.integers(max(18, days // 15), max(30, days // 4)))
    elif pattern == "seasonal":
        count = int(rng.integers(max(12, days // 20), max(24, days // 8)))
    else:
        count = int(rng.integers(max(6, days // 50), max(12, days // 18)))

    offsets = rng.integers(0, days + 1, size=count)
    dates = [start_date + timedelta(days=int(offset)) for offset in offsets]
    dates.sort()
    return dates


def quantity_for_pattern(rng: np.random.Generator, pattern: str, current_date: date, base_monthly_demand: float, decan: int) -> float:
    if pattern == "zero":
        return 0.0

    daily_base = max(base_monthly_demand / 4.0, 1.0)
    if pattern == "stable":
        quantity = rng.normal(daily_base, daily_base * 0.15)
    elif pattern == "volatile":
        quantity = rng.normal(daily_base, daily_base * 0.70)
    elif pattern == "seasonal":
        seasonal_factor = 1.0 + 0.60 * math.sin((current_date.timetuple().tm_yday / 365.0) * 2 * math.pi)
        quantity = rng.normal(daily_base * seasonal_factor, daily_base * 0.25)
    else:
        burst_multiplier = 1.0 if rng.random() > 0.35 else rng.uniform(2.0, 4.5)
        quantity = rng.normal(daily_base * burst_multiplier, daily_base * 0.85)

    quantity = max(quantity, 0.0)
    if decan == 0:
        return float(max(1, int(round(quantity))))
    return round(float(max(quantity, 0.001)), decan)


def calculate_safety_stock(demand_quantities: list[float], lead_time_days: int, abc_class: str, decan: int) -> float:
    if not demand_quantities:
        return 0.0
    if len(demand_quantities) > 1:
        demand_std = float(np.std(demand_quantities, ddof=1))
    else:
        demand_std = float(demand_quantities[0]) * 0.1
    safety_stock = Z_VALUES[abc_class] * demand_std * math.sqrt(lead_time_days)
    if decan == 0:
        return float(max(0, int(math.ceil(safety_stock))))
    return round(float(max(0.0, safety_stock)), decan)


def build_materials(rng: np.random.Generator, plants: list[str], materials_per_plant: int, start_date: date, end_date: date) -> tuple[list[MaterialRecord], pd.DataFrame]:
    materials: list[MaterialRecord] = []
    matdoc_rows: list[dict[str, object]] = []
    material_doc_number = 5000000000

    for plant_idx, werks in enumerate(plants):
        for material_idx in range(materials_per_plant):
            matnr = generate_material_id(plant_idx, material_idx)
            mtart = str(rng.choice(["ROH", "HALB", "FERT"], p=[0.45, 0.25, 0.30]))
            meins = str(rng.choice(list(UOM_CONFIG.keys()), p=[0.45, 0.30, 0.15, 0.10]))
            decan = UOM_CONFIG[meins]["decan"]
            abc_class = weighted_choice(rng, ABC_DISTRIBUTION)
            demand_pattern = weighted_choice(rng, DEMAND_PATTERNS)
            lead_time_days = int(rng.integers(5, 61))
            goods_receipt_days = int(rng.integers(1, 8))
            purchasing_processing_days = 0
            price_control = weighted_choice(rng, PRICE_CONTROL_DISTRIBUTION)
            base_price = generate_price(rng)
            standard_price = base_price if price_control == "S" else round(base_price * rng.uniform(0.95, 1.15), 2)
            moving_average_price = base_price if price_control == "V" else round(base_price * rng.uniform(0.90, 1.10), 2)
            base_monthly_demand = float(rng.uniform(20, 400))

            dates = generate_demand_dates(rng, demand_pattern, start_date, end_date)
            issued_quantities: list[float] = []
            for movement_date in dates:
                quantity = quantity_for_pattern(rng, demand_pattern, movement_date, base_monthly_demand, decan)
                if quantity <= 0:
                    continue
                shkzg = str(rng.choice(MOVEMENT_SIGNS, p=[0.92, 0.08]))
                bwart = int(rng.choice(MOVEMENT_TYPES))
                material_doc_number += 1
                matdoc_rows.append(
                    {
                        "MATNR": matnr,
                        "WERKS": werks,
                        "BWART": bwart,
                        "SHKZG": shkzg,
                        "MENGE": quantity if decan > 0 else int(quantity),
                        "BUDAT": pd.Timestamp(movement_date),
                        "MJAHR": str(movement_date.year),
                        "MBLNR": str(material_doc_number),
                        "ZEILE": "0001",
                        "SJAHR": "",
                        "SMBLN": "",
                        "SMBLP": "",
                    }
                )
                if shkzg == "S":
                    issued_quantities.append(float(quantity))

            safety_stock = calculate_safety_stock(issued_quantities, lead_time_days, abc_class, decan)
            wzeit_days = None
            if rng.random() < 0.30:
                wzeit_days = int(round(lead_time_days + (5.0 / 7.0) * (purchasing_processing_days + goods_receipt_days)))
            materials.append(
                MaterialRecord(
                    matnr=matnr,
                    werks=werks,
                    mtart=mtart,
                    meins=meins,
                    demand_pattern=demand_pattern,
                    abc_class=abc_class,
                    lead_time_days=lead_time_days,
                    goods_receipt_days=goods_receipt_days,
                    purchasing_processing_days=purchasing_processing_days,
                    wzeit_days=wzeit_days,
                    safety_stock=safety_stock,
                    price_control=price_control,
                    standard_price=standard_price,
                    moving_average_price=moving_average_price,
                    deletion_flag="",
                    plant_deletion_flag="",
                    base_monthly_demand=base_monthly_demand,
                )
            )

    apply_special_safety_stock_rules(rng, materials)
    matdoc_df = pd.DataFrame(matdoc_rows)
    if matdoc_df.empty:
        matdoc_df = pd.DataFrame(columns=[tech for _, tech in TEMPLATE_STRUCTURE["MATDOC"]])
    else:
        matdoc_df = append_storno_rows(rng, matdoc_df, material_doc_number)
        matdoc_df = matdoc_df.sort_values(["WERKS", "MATNR", "BUDAT", "MBLNR"]).reset_index(drop=True)
    return materials, matdoc_df


def append_storno_rows(rng: np.random.Generator, matdoc_df: pd.DataFrame, last_material_doc_number: int) -> pd.DataFrame:
    if matdoc_df.empty:
        return matdoc_df

    eligible = matdoc_df[
        matdoc_df["SJAHR"].astype(str).str.strip().eq("")
        & matdoc_df["SMBLN"].astype(str).str.strip().eq("")
        & matdoc_df["SMBLP"].astype(str).str.strip().eq("")
    ].copy()
    if eligible.empty:
        return matdoc_df

    storno_count = max(1, int(round(len(eligible) * 0.02)))
    storno_count = min(storno_count, len(eligible))
    selected_indices = rng.choice(eligible.index.to_numpy(), size=storno_count, replace=False)

    storno_rows: list[dict[str, object]] = []
    next_doc_number = last_material_doc_number
    for idx in selected_indices:
        original = matdoc_df.loc[idx]
        next_doc_number += 1
        original_date = pd.Timestamp(original["BUDAT"])
        reversal_date = original_date + pd.Timedelta(days=int(rng.integers(1, 15)))
        if reversal_date.date() > datetime.utcnow().date():
            reversal_date = pd.Timestamp(datetime.utcnow().date())

        storno_rows.append(
            {
                "MATNR": original["MATNR"],
                "WERKS": original["WERKS"],
                "BWART": original["BWART"],
                "SHKZG": "H" if str(original["SHKZG"]).upper() == "S" else "S",
                "MENGE": original["MENGE"],
                "BUDAT": reversal_date,
                "MJAHR": str(reversal_date.year),
                "MBLNR": str(next_doc_number),
                "ZEILE": "0001",
                "SJAHR": str(original["MJAHR"]),
                "SMBLN": str(original["MBLNR"]),
                "SMBLP": str(original["ZEILE"]),
            }
        )

    if not storno_rows:
        return matdoc_df
    return pd.concat([matdoc_df, pd.DataFrame(storno_rows)], ignore_index=True)


def apply_special_safety_stock_rules(rng: np.random.Generator, materials: list[MaterialRecord]) -> None:
    total = len(materials)
    if total == 0:
        return

    null_count = max(1, int(round(total * 0.10)))
    high_count = max(1, int(round(total * 0.15)))
    indices = np.arange(total)
    rng.shuffle(indices)
    null_indices = set(indices[:null_count].tolist())
    high_indices = set(indices[null_count:null_count + high_count].tolist())

    for idx, material in enumerate(materials):
        if idx in null_indices:
            materials[idx] = MaterialRecord(**{**material.__dict__, "safety_stock": None})
        elif idx in high_indices and material.safety_stock is not None:
            multiplier = float(rng.uniform(2.0, 4.0))
            bumped = material.safety_stock * multiplier
            decan = UOM_CONFIG[material.meins]["decan"]
            if decan == 0:
                bumped = float(int(math.ceil(bumped)))
            else:
                bumped = round(float(bumped), decan)
            materials[idx] = MaterialRecord(**{**material.__dict__, "safety_stock": bumped})


def build_tables(rng: np.random.Generator, materials: list[MaterialRecord], plants: list[str], matdoc: pd.DataFrame) -> dict[str, pd.DataFrame]:
    abc_visibility_mask = np.ones(len(materials), dtype=bool)
    hidden_count = len(materials) // 2
    if hidden_count:
        hidden_indices = rng.choice(len(materials), size=hidden_count, replace=False)
        abc_visibility_mask[hidden_indices] = False

    t001 = pd.DataFrame([{"BUKRS": company_code(werks), "WAERS": "EUR"} for werks in plants]).drop_duplicates(subset=["BUKRS"]).reset_index(drop=True)
    t001w = pd.DataFrame([{"WERKS": werks, "BWKEY": werks} for werks in plants])
    t006 = pd.DataFrame([{"MSEHI": uom, "DECAN": config["decan"]} for uom, config in UOM_CONFIG.items()])
    t134 = pd.DataFrame([
        {"MTART": "ROH", "PROD_TYPE_CODE": 1},
        {"MTART": "HALB", "PROD_TYPE_CODE": 1},
        {"MTART": "FERT", "PROD_TYPE_CODE": 1},
    ])
    v438a = pd.DataFrame([
        {"DISMM": "PD", "DISVF": "Y"},
        {"DISMM": "VB", "DISVF": "Y"},
    ])
    if len(plants) <= 1:
        bztek_map = {plants[0]: 0} if plants else {}
    else:
        bztek_map = {plants[0]: 0}
        non_zero_assigned = False
        for werks in plants[1:]:
            value = int(rng.integers(0, 3))
            if value > 0:
                non_zero_assigned = True
            bztek_map[werks] = value
        if not non_zero_assigned and len(plants) > 1:
            chosen_plant = str(rng.choice(plants[1:]))
            bztek_map[chosen_plant] = int(rng.integers(1, 3))

    v399 = pd.DataFrame(
        [{"WERKS": werks, "BZTEK": bztek_map.get(werks, 0)} for werks in plants]
    )
    mara = pd.DataFrame([
        {"MATNR": material.matnr, "MTART": material.mtart, "LVORM": material.deletion_flag, "MEINS": material.meins}
        for material in materials
    ]).drop_duplicates(subset=["MATNR"]).reset_index(drop=True)

    marc_rows = []
    for idx, material in enumerate(materials):
        maabc = material.abc_class if abc_visibility_mask[idx] else ""
        marc_rows.append(
            {
                "MATNR": material.matnr,
                "WERKS": material.werks,
                "LVORM": material.plant_deletion_flag,
                "DISMM": "PD",
                "EISBE": material.safety_stock,
                "PLIFZ": material.lead_time_days,
                "WEBAZ": material.goods_receipt_days,
                "PSTAT": "DEL",
                "MAABC": maabc,
                "WZEIT": material.wzeit_days,
            }
        )
    marc = pd.DataFrame(marc_rows)

    mbew = pd.DataFrame([
        {
            "MATNR": material.matnr,
            "BWKEY": material.werks,
            "VPRSV": material.price_control,
            "STPRS": material.standard_price,
            "VERPR": material.moving_average_price,
            "PEINH": 1,
            "LVORM": "",
        }
        for material in materials
    ])

    tables = {
        "T001": t001,
        "T001W": t001w,
        "T006": t006,
        "T134": t134,
        "V438A": v438a,
        "V_399D_E": v399,
        "MARA": mara,
        "MATDOC": matdoc,
        "MARC": marc,
        "MBEW": mbew,
    }

    ordered_tables: dict[str, pd.DataFrame] = {}
    for sheet_name, columns in TEMPLATE_STRUCTURE.items():
        ordered_tables[sheet_name] = tables[sheet_name][[tech for _, tech in columns]].copy()
    return ordered_tables


def generate_sap_dataset(num_plants: int = 4, materials_per_plant: int = 100, years_of_history: int = 3) -> dict[str, object]:
    rng, seed = build_rng()
    end_date = datetime.utcnow().date()
    start_date = end_date - timedelta(days=365 * years_of_history)
    plants = generate_plant_codes(num_plants)
    materials, matdoc = build_materials(rng, plants, materials_per_plant, start_date, end_date)
    tables = build_tables(rng, materials, plants, matdoc)

    stats = {
        "total_materials": len(tables["MARC"]),
        "total_plants": len(plants),
        "matdoc_rows": len(tables["MATDOC"]),
    }
    metadata = {
        "seed": seed,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
    }
    return {"tables": tables, "stats": stats, "metadata": metadata}


def create_excel_file(tables: dict[str, pd.DataFrame]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, columns in TEMPLATE_STRUCTURE.items():
            df = tables[sheet_name]
            df.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=2)
            ws = writer.book[sheet_name]

            for col_idx, (business_name, technical_name) in enumerate(columns, start=1):
                ws.cell(row=1, column=col_idx, value=business_name)
                header_cell = ws.cell(row=2, column=col_idx, value=technical_name)
                header_cell.font = Font(bold=True)

            last_col = get_column_letter(len(columns))
            last_row = max(len(df) + 2, 2)
            ws.auto_filter.ref = f"A2:{last_col}{last_row}"

            for col_idx in range(1, len(columns) + 1):
                max_len = 0
                for row_idx in range(1, min(ws.max_row, 200) + 1):
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if value is None:
                        continue
                    max_len = max(max_len, len(str(value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 42)

    output.seek(0)
    return output.getvalue()
